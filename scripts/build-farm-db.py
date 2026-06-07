from __future__ import annotations

import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "瑪奇農場模型(ver.241123).xlsx"
OUTPUT = ROOT / "data" / "farm-db.js"

FEATURED_SHEETS = ["素質", "素質(額外)", "素質(小屋)", "素質(女神像)"]
META_HINTS = {"台服", "外服", "來源", "備註", "註記", "說明"}


def clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def is_meaningful(value: Any) -> bool:
    if value in (None, ""):
        return False
    if isinstance(value, (int, float)):
      return value != 0
    return True


def normalize_headers(rows: list[tuple[Any, ...]]) -> tuple[list[str], list[int]]:
    first_row = list(rows[0])
    data_rows = rows[1:]
    active_indexes: list[int] = []

    for idx in range(len(first_row)):
        header = clean_value(first_row[idx]) if idx < len(first_row) else ""
        has_data = any(is_meaningful(clean_value(row[idx])) for row in data_rows if idx < len(row))
        if is_meaningful(header) or has_data:
            active_indexes.append(idx)

    names: list[str] = []
    blanks = 0
    seen = Counter()

    for idx in active_indexes:
        raw = clean_value(first_row[idx])
        if not raw:
            blanks += 1
            raw = f"備註{blanks}"
        name = str(raw)
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        names.append(name)

    return names, active_indexes


def detect_ability_fields(headers: list[str], rows: list[dict[str, Any]]) -> list[str]:
    results: list[str] = []
    meta_keywords = tuple(META_HINTS)

    for header in headers:
        if header == "名稱" or any(keyword in header for keyword in meta_keywords):
            continue

        values = [row.get(header) for row in rows if is_meaningful(row.get(header))]
        if not values:
            continue

        numeric_values = [value for value in values if isinstance(value, (int, float))]
        ratio = len(numeric_values) / len(values)

        if len(numeric_values) >= 2 and ratio >= 0.6:
            results.append(header)

    return results


def build_sheet(sheet) -> dict[str, Any]:
    raw_rows = list(sheet.iter_rows(values_only=True))
    headers, indexes = normalize_headers(raw_rows)

    rows: list[dict[str, Any]] = []
    for row_index, raw in enumerate(raw_rows[1:], start=2):
        values: dict[str, Any] = {}
        for header, idx in zip(headers, indexes):
            value = clean_value(raw[idx] if idx < len(raw) else "")
            if is_meaningful(value):
                values[header] = value

        if not values:
            continue

        name = str(values.get("名稱") or values.get(headers[0]) or f"{sheet.title}-{row_index}")
        rows.append(values | {"名稱": name})

    ability_fields = detect_ability_fields(headers, rows)
    meta_fields = [
        header
        for header in headers
        if header != "名稱" and header not in ability_fields
    ]

    normalized_rows = []
    for row_number, row in enumerate(rows, start=1):
        stats = {
            header: row[header]
            for header in ability_fields
            if is_meaningful(row.get(header))
        }
        meta = {
            header: row[header]
            for header in meta_fields
            if is_meaningful(row.get(header))
        }
        normalized_rows.append(
            {
                "id": f"{sheet.title}-{row_number}",
                "name": row["名稱"],
                "values": {header: row.get(header, "") for header in headers},
                "stats": stats,
                "meta": meta,
            }
        )

    return {
        "name": sheet.title,
        "fields": headers,
        "abilityFields": ability_fields,
        "metaFields": meta_fields,
        "rows": normalized_rows,
    }


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(SOURCE, data_only=True)

    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sourceFile": SOURCE.name,
        "featuredSheets": [name for name in FEATURED_SHEETS if name in workbook.sheetnames],
        "sheets": [build_sheet(workbook[name]) for name in workbook.sheetnames],
    }

    js = "window.FARM_DB = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUTPUT.write_text(js, encoding="utf-8")
    print(f"Built {OUTPUT}")


if __name__ == "__main__":
    main()
